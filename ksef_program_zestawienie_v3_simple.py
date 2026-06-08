import os
import re
import traceback
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright


class KsefSimpleSummaryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("KSeF Program do Arkuszy")
        self.root.geometry("1180x760")
        self.root.minsize(1100, 720)
        self.root.configure(bg="#eef2f7")

        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None

        self.logo_image = None
        self.spinner_frames = ["●○○", "○●○", "○○●", "○●○"]
        self.spinner_index = 0
        self.is_busy = False

        self.base_dir = os.getcwd()
        self.output_dir = os.path.join(self.base_dir, "zestawienia_ksef")
        os.makedirs(self.output_dir, exist_ok=True)

        self.pages_var = tk.StringVar(value="0")
        self.rows_var = tk.StringVar(value="0")
        self.file_var = tk.StringVar(value="Brak")
        self.status_var = tk.StringVar(value="Gotowe do pracy")
        self.spinner_var = tk.StringVar(value="○○○")

        self.status_box = None
        self.progress = None

        self.build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def build_ui(self):
        main = tk.Frame(self.root, bg="#eef2f7")
        main.pack(fill="both", expand=True, padx=18, pady=18)

        header = tk.Frame(main, bg="#ffffff", bd=1, relief="solid")
        header.pack(fill="x")

        accent = tk.Frame(header, bg="#c81f25", height=5)
        accent.pack(fill="x")
        accent.pack_propagate(False)

        header_inner = tk.Frame(header, bg="#ffffff")
        header_inner.pack(fill="x", padx=18, pady=16)

        left_head = tk.Frame(header_inner, bg="#ffffff")
        left_head.pack(side="left", fill="y")

        self._load_logo(left_head)

        title_wrap = tk.Frame(left_head, bg="#ffffff")
        title_wrap.pack(side="left", padx=(14, 0))

        tk.Label(
            title_wrap,
            text="KSeF Program do Arkuszy",
            bg="#ffffff",
            fg="#0f172a",
            font=("Segoe UI", 23, "bold"),
        ).pack(anchor="w")

        tk.Label(
            title_wrap,
            text="by Paweł Ruchlicki",
            bg="#ffffff",
            fg="#64748b",
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w", pady=(6, 0))

        right_head = tk.Frame(header_inner, bg="#ffffff")
        right_head.pack(side="right", anchor="e")

        status_pill = tk.Frame(right_head, bg="#f8fafc", bd=1, relief="solid")
        status_pill.pack(anchor="e")

        tk.Label(
            status_pill,
            textvariable=self.spinner_var,
            bg="#f8fafc",
            fg="#c81f25",
            font=("Segoe UI", 11, "bold"),
            padx=10,
            pady=7,
        ).pack(side="left")

        tk.Label(
            status_pill,
            textvariable=self.status_var,
            bg="#f8fafc",
            fg="#0f172a",
            font=("Segoe UI", 10, "bold"),
            padx=2,
            pady=7,
        ).pack(side="left")

        content = tk.Frame(main, bg="#eef2f7")
        content.pack(fill="both", expand=True, pady=(14, 0))

        left = tk.Frame(content, bg="#ffffff", bd=1, relief="solid", padx=16, pady=16)
        left.pack(side="left", fill="y", padx=(0, 12))
        left.configure(width=355)
        left.pack_propagate(False)

        right = tk.Frame(content, bg="#ffffff", bd=1, relief="solid", padx=16, pady=16)
        right.pack(side="left", fill="both", expand=True)

        tk.Label(left, text="Sterowanie", font=("Segoe UI", 16, "bold"), bg="#ffffff", fg="#0f172a").pack(anchor="w")

        btn_wrap = tk.Frame(left, bg="#ffffff")
        btn_wrap.pack(fill="x", pady=(12, 8))

        self.make_btn(btn_wrap, "Otwórz KSeF", self.open_ksef, primary=True).pack(fill="x", pady=(0, 10))
        self.make_btn(btn_wrap, "Pobierz zestawienie do Excel", self.export_summary, primary=False).pack(fill="x", pady=(0, 10))
        self.make_btn(btn_wrap, "Otwórz folder z plikami", self.open_output_folder, primary=False).pack(fill="x")

        stats_wrap = tk.Frame(left, bg="#ffffff")
        stats_wrap.pack(fill="x", pady=(14, 14))

        self.make_stat(stats_wrap, "Strony", self.pages_var).pack(fill="x", pady=(0, 8))
        self.make_stat(stats_wrap, "Wiersze", self.rows_var).pack(fill="x", pady=(0, 8))
        self.make_stat(stats_wrap, "Plik", self.file_var, small=True).pack(fill="x")

        footer = tk.Frame(left, bg="#ffffff")
        footer.pack(side="bottom", fill="x", pady=(12, 0))
        tk.Label(
            footer,
            text="by Paweł Ruchlicki",
            bg="#ffffff",
            fg="#64748b",
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w")

        tk.Label(right, text="Log operacji", font=("Segoe UI", 16, "bold"), bg="#ffffff", fg="#0f172a").pack(anchor="w")

        self.progress = ttk.Progressbar(right, mode="determinate", maximum=100)
        self.progress.pack(fill="x", pady=(12, 12))

        log_frame = tk.Frame(right, bg="#0b1220")
        log_frame.pack(fill="both", expand=True)

        self.status_box = tk.Text(
            log_frame,
            bg="#0b1220",
            fg="#dbe7f4",
            insertbackground="#ffffff",
            relief="flat",
            bd=0,
            font=("Consolas", 10),
            wrap="word",
            padx=12,
            pady=12,
        )
        self.status_box.pack(fill="both", expand=True)

        self.log("[INFO] Program uruchomiony.")
        self.log("[INFO] Filtry ustawiasz ręcznie w KSeF.")
        self.log("[INFO] Program zapisze dokładnie to, co jest aktualnie widoczne na liście.")
        self.progress["value"] = 0

    def make_btn(self, parent, text, command, primary=False):
        bg = "#c81f25" if primary else "#ffffff"
        fg = "#ffffff" if primary else "#0f172a"
        active = "#a8181d" if primary else "#f1f5f9"
        return tk.Button(
            parent,
            text=text,
            command=command,
            bg=bg,
            fg=fg,
            activebackground=active,
            activeforeground=fg,
            relief="flat",
            bd=0,
            font=("Segoe UI", 10, "bold"),
            cursor="hand2",
            padx=14,
            pady=12,
        )

    def make_stat(self, parent, label, var, small=False):
        card = tk.Frame(parent, bg="#f8fafc", bd=1, relief="solid")
        tk.Label(card, text=label, bg="#f8fafc", fg="#64748b", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=12, pady=(10, 2))
        tk.Label(
            card,
            textvariable=var,
            bg="#f8fafc",
            fg="#0f172a",
            font=("Segoe UI", 10 if small else 18, "bold"),
            wraplength=280,
            justify="left",
        ).pack(anchor="w", padx=12, pady=(0, 10))
        return card

    def _load_logo(self, parent):
        logo_wrap = tk.Frame(parent, bg="#ffffff", width=78, height=78)
        logo_wrap.pack(side="left")
        logo_wrap.pack_propagate(False)

        logo_path = self.find_logo_path()
        if logo_path:
            try:
                self.logo_image = tk.PhotoImage(file=logo_path)
                tk.Label(logo_wrap, image=self.logo_image, bg="#ffffff").pack(fill="both", expand=True)
                return
            except Exception:
                pass

        tk.Label(
            logo_wrap,
            text="KSeF",
            bg="#c81f25",
            fg="#ffffff",
            font=("Segoe UI", 16, "bold"),
        ).pack(fill="both", expand=True)

    def find_logo_path(self):
        graphics_dir = os.path.join(self.base_dir, "Grafiki")
        candidates = [
            os.path.join(graphics_dir, "logo.png"),
            os.path.join(graphics_dir, "logo.gif"),
            os.path.join(graphics_dir, "logo.jpg"),
            os.path.join(graphics_dir, "logo.jpeg"),
        ]
        for path in candidates:
            if os.path.exists(path):
                return path
        if os.path.isdir(graphics_dir):
            for name in os.listdir(graphics_dir):
                if name.lower().endswith((".png", ".gif")):
                    return os.path.join(graphics_dir, name)
        return None

    def ui_tick(self):
        self.spinner_index = (self.spinner_index + 1) % len(self.spinner_frames)
        self.spinner_var.set(self.spinner_frames[self.spinner_index] if self.is_busy else "○○○")
        self.root.update_idletasks()
        self.root.update()

    def set_busy(self, value, status=None):
        self.is_busy = value
        if status is not None:
            self.status_var.set(status)
        self.ui_tick()

    def set_status(self, text):
        self.status_var.set(text)
        self.ui_tick()

    def log(self, text):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.status_box.insert("end", f"[{timestamp}] {text}\n")
        self.status_box.see("end")
        self.ui_tick()

    def set_progress(self, current, total):
        total = max(1, total)
        pct = int((current / total) * 100)
        self.progress["value"] = pct
        self.ui_tick()

    def open_output_folder(self):
        os.makedirs(self.output_dir, exist_ok=True)
        try:
            os.startfile(self.output_dir)
        except Exception:
            subprocess.Popen(["explorer", self.output_dir])

    def open_ksef(self):
        if self.page is not None:
            messagebox.showinfo("Informacja", "Okno KSeF jest już otwarte.")
            return
        try:
            self.set_busy(True, "Uruchamianie KSeF...")
            self.log("[INFO] Uruchamiam przeglądarkę...")
            self.playwright = sync_playwright().start()
            try:
                self.browser = self.playwright.chromium.launch(channel="msedge", headless=False, slow_mo=70)
            except Exception:
                self.browser = self.playwright.chromium.launch(headless=False, slow_mo=70)
            self.context = self.browser.new_context(accept_downloads=True)
            self.page = self.context.new_page()
            self.page.goto("https://ap.ksef.mf.gov.pl/web/invoice-list", timeout=90000)
            self.set_busy(False, "KSeF otwarty")
            self.log("[OK] KSeF został otwarty.")
            self.log("[INFO] Zaloguj się ręcznie, ustaw filtry ręcznie i przejdź na listę faktur.")
            messagebox.showinfo(
                "KSeF",
                "KSeF został otwarty.\n\n"
                "1. Zaloguj się ręcznie\n"
                "2. Ustaw filtry ręcznie\n"
                "3. Otwórz listę faktur zakupu\n"
                "4. Kliknij „Pobierz zestawienie do Excel”",
            )
        except Exception as e:
            self.set_busy(False, "Błąd")
            self.log(f"[BŁĄD] Nie udało się otworzyć KSeF: {e}")
            messagebox.showerror("Błąd", f"Nie udało się otworzyć KSeF.\n\n{e}")

    @staticmethod
    def normalize(text):
        return re.sub(r"\s+", " ", (text or "")).strip()

    @staticmethod
    def is_money_cell(text):
        return bool(re.fullmatch(r"[-+]?\d[\d\s\.]*,\d{2}\s+[A-Z]{3}", text or ""))

    @staticmethod
    def clean_invoice_number(text):
        text = text or ""
        text = re.sub(r"\s*Kopiuj numer faktury.*$", "", text, flags=re.I)
        text = re.sub(r"\s*Kopiuj numer KSeF.*$", "", text, flags=re.I)
        text = re.sub(r"\s*content_copy.*$", "", text, flags=re.I)
        text = re.sub(r"\s*Przejdź do podglądu faktury.*$", "", text, flags=re.I)
        text = re.sub(r"\s*\(online\)\s*$", "", text, flags=re.I)
        text = re.sub(r"\s+", " ", text).strip()
        if text.lower() in {"online", "(online)", "podgląd", "podglad"}:
            return ""
        return text.strip()

    @staticmethod
    def clean_date(text):
        m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", text or "")
        if m:
            d, mth, y = m.groups()
            return f"{y}-{mth}-{d}"
        return ""

    @staticmethod
    def parse_money(text):
        m = re.search(r"([-+]?\d[\d\s\.]*,\d{2})\s+([A-Z]{3})", text or "")
        if not m:
            return "", None
        amount_txt, currency = m.groups()
        amount = float(amount_txt.replace(" ", "").replace(".", "").replace(",", "."))
        return currency, amount

    @staticmethod
    def extract_nr_ksef(text):
        m = re.search(r"\b\d{10}-\d{8}-[A-Z0-9]+-[A-Z0-9]+\b", text or "", flags=re.I)
        return m.group(0) if m else ""


    def is_noise_text(self, text):
        low = self.normalize(text).lower()
        if not low:
            return True
        noise_parts = [
            "przejdź do podglądu faktury",
            "kopiuj numer faktury",
            "kopiuj numer ksef",
            "content_copy",
        ]
        if any(part in low for part in noise_parts):
            return True
        if low in {"online", "(online)", "podgląd", "podglad", "wybierz", "akcje"}:
            return True
        return False

    def parse_row_by_position(self, filtered):
        if len(filtered) < 10:
            return None

        identifier = filtered[0] if re.fullmatch(r"\d{10}", filtered[0]) else ""
        seller_name = filtered[1] if len(filtered) > 1 else ""
        nr_ksef = self.extract_nr_ksef(filtered[2]) if len(filtered) > 2 else ""
        nr_faktury = self.clean_invoice_number(filtered[3]) if len(filtered) > 3 else ""

        if self.is_noise_text(nr_faktury):
            nr_faktury = ""

        issue_date = self.clean_date(filtered[4]) if len(filtered) > 4 else ""
        saved_date = self.clean_date(filtered[5]) if len(filtered) > 5 else ""
        received_date = self.clean_date(filtered[6]) if len(filtered) > 6 else ""

        waluta = ""
        netto = brutto = vat_pln = None
        if len(filtered) > 7:
            waluta, netto = self.parse_money(filtered[7])
        if len(filtered) > 8:
            _, brutto = self.parse_money(filtered[8])
        if len(filtered) > 9:
            _, vat_pln = self.parse_money(filtered[9])

        if not nr_ksef and not nr_faktury:
            return None

        return {
            "Identyfikator sprzedawcy": identifier,
            "Nazwa sprzedawcy": seller_name,
            "Nr KSeF": nr_ksef,
            "Nr faktury": nr_faktury,
            "Data wystawienia": issue_date,
            "Data zapisania w KSeF": saved_date,
            "Data otrzymania": received_date,
            "Waluta": waluta,
            "Netto": netto,
            "Brutto": brutto,
            "VAT (PLN)": vat_pln,
            "_signature": f"{identifier}|{nr_ksef}|{nr_faktury}",
        }

    def get_rows_on_page(self):
        selectors = ["table tbody tr", "tbody tr", "[role='row']"]
        rows = None
        for sel in selectors:
            try:
                loc = self.page.locator(sel)
                if loc.count() > 0:
                    rows = loc
                    break
            except Exception:
                continue

        parsed = []
        if rows is None:
            return parsed

        count = rows.count()
        for i in range(count):
            try:
                row = rows.nth(i)
                if not row.is_visible():
                    continue

                cell_loc = row.locator("td, [role='cell']")
                if cell_loc.count() == 0:
                    continue

                cells = []
                for c in range(cell_loc.count()):
                    try:
                        txt = self.normalize(cell_loc.nth(c).inner_text())
                    except Exception:
                        txt = ""
                    if txt:
                        cells.append(txt)

                if not cells:
                    continue

                item = self.parse_row_cells(cells)
                if item:
                    parsed.append(item)
            except Exception:
                continue

        return parsed

    def parse_row_cells(self, cells):
        filtered = []
        for txt in cells:
            t = self.normalize(txt)
            if not t:
                continue
            if "zaznacz tylko ten wiersz" in t.lower():
                continue
            if t.lower() in {"wybierz", "akcje"}:
                continue
            filtered.append(t)

        if len(filtered) < 6:
            return None

        # Najpierw próbujemy mapowania po stałej pozycji kolumn.
        # To zabezpiecza przypadki, gdy numer FV wygląda jak data, np. 1.12.2025.
        positional = self.parse_row_by_position(filtered)
        if positional:
            return positional

        identifier = ""
        seller_name = ""
        nr_ksef = ""
        nr_faktury = ""
        dates = []
        amount_cells = []

        for t in filtered:
            if not identifier and re.fullmatch(r"\d{10}", t):
                identifier = t
                continue

            if not nr_ksef:
                maybe = self.extract_nr_ksef(t)
                if maybe:
                    nr_ksef = maybe
                    continue

            if self.clean_date(t):
                dates.append(self.clean_date(t))
                continue

            if self.is_money_cell(t):
                amount_cells.append(t)
                continue

        if identifier:
            try:
                idx = filtered.index(identifier)
                if idx + 1 < len(filtered):
                    seller_name = filtered[idx + 1]
            except Exception:
                pass

        ksef_index = -1
        for idx, t in enumerate(filtered):
            if nr_ksef and nr_ksef in t:
                ksef_index = idx
                break

        if ksef_index >= 0:
            for t in filtered[ksef_index + 1:]:
                if self.is_money_cell(t):
                    continue
                if re.fullmatch(r"\d{10}", t):
                    continue
                if self.extract_nr_ksef(t):
                    continue
                if t == seller_name:
                    continue
                candidate = self.clean_invoice_number(t)
                if self.is_noise_text(candidate):
                    continue
                nr_faktury = candidate
                if nr_faktury:
                    break

        if not seller_name:
            for t in filtered:
                if t == identifier:
                    continue
                if self.extract_nr_ksef(t) or self.is_money_cell(t) or self.clean_date(t):
                    continue
                if self.is_noise_text(t):
                    continue
                seller_name = t
                break

        if not nr_ksef and not nr_faktury and not dates:
            return None

        netto_currency = ""
        netto = brutto = vat_pln = None

        if len(amount_cells) >= 1:
            netto_currency, netto = self.parse_money(amount_cells[0])
        if len(amount_cells) >= 2:
            _, brutto = self.parse_money(amount_cells[1])
        if len(amount_cells) >= 3:
            _, vat_pln = self.parse_money(amount_cells[2])

        return {
            "Identyfikator sprzedawcy": identifier,
            "Nazwa sprzedawcy": seller_name,
            "Nr KSeF": nr_ksef,
            "Nr faktury": nr_faktury,
            "Data wystawienia": dates[0] if len(dates) > 0 else "",
            "Data zapisania w KSeF": dates[1] if len(dates) > 1 else "",
            "Data otrzymania": dates[2] if len(dates) > 2 else "",
            "Waluta": netto_currency,
            "Netto": netto,
            "Brutto": brutto,
            "VAT (PLN)": vat_pln,
            "_signature": f"{identifier}|{nr_ksef}|{nr_faktury}",
        }

    def page_signature(self, rows):
        return "|".join(r["_signature"] for r in rows[:5]) if rows else "EMPTY"

    def is_disabled_element(self, locator):
        try:
            disabled = (locator.get_attribute("disabled") or "") + (locator.get_attribute("aria-disabled") or "")
            if "true" in disabled.lower():
                return True
            classes = (locator.get_attribute("class") or "").lower()
            if "disabled" in classes:
                return True
        except Exception:
            return False
        return False

    def click_locator(self, locator, timeout=3000):
        try:
            locator.scroll_into_view_if_needed(timeout=timeout)
        except Exception:
            pass
        for kwargs in ({}, {"force": True}):
            try:
                locator.click(timeout=timeout, **kwargs)
                return True
            except Exception:
                pass
        try:
            locator.evaluate("el => el.click()")
            return True
        except Exception:
            return False

    def current_numeric_page(self):
        selectors = [
            "[aria-current='page']",
            "button[aria-current='true']",
            "a[aria-current='true']",
            ".active",
            ".selected",
            ".current",
        ]
        for sel in selectors:
            try:
                loc = self.page.locator(sel)
                for i in range(loc.count()):
                    txt = self.normalize(loc.nth(i).inner_text())
                    if txt.isdigit():
                        return int(txt)
            except Exception:
                continue
        return None

    def wait_for_page_change(self, before_signature, before_page=None, timeout_ms=7000):
        loops = max(1, timeout_ms // 250)
        for _ in range(loops):
            self.page.wait_for_timeout(250)
            rows = self.get_rows_on_page()
            after_signature = self.page_signature(rows)
            after_page = self.current_numeric_page()
            if after_signature and after_signature != before_signature:
                return True
            if before_page is not None and after_page is not None and after_page != before_page:
                return True
        return False

    def ensure_invoice_list_view(self):
        try:
            url = self.page.url.lower()
        except Exception:
            return
        if "invoice-list" in url:
            return
        self.log(f"[INFO] Wykryto wyjście z listy ({url}). Wracam do listy.")
        try:
            self.page.go_back(timeout=10000)
            self.page.wait_for_timeout(1200)
        except Exception:
            pass

    def scroll_to_pagination_area(self):
        try:
            self.page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            self.page.wait_for_timeout(500)
        except Exception:
            pass

    def pagination_selectors(self):
        return [
            "[class*='pagination'] button",
            "[class*='pagination'] a",
            "[class*='pager'] button",
            "[class*='pager'] a",
            "nav button",
            "nav a",
            "[aria-label*='page']",
            "[title*='page']",
            "[aria-label*='strona']",
            "[title*='strona']",
        ]

    def click_page_number(self, target_page):
        if target_page is None:
            return False
        self.scroll_to_pagination_area()
        before_signature = self.page_signature(self.get_rows_on_page())
        before_page = self.current_numeric_page()

        selectors = self.pagination_selectors()
        for sel in selectors:
            try:
                loc = self.page.locator(sel)
                if loc.count() == 0:
                    continue
                for i in range(loc.count()):
                    item = loc.nth(i)
                    txt = self.normalize(item.inner_text())
                    if txt != str(target_page):
                        continue
                    if self.is_disabled_element(item):
                        continue
                    if not self.click_locator(item, timeout=2500):
                        continue
                    self.ensure_invoice_list_view()
                    if self.wait_for_page_change(before_signature, before_page, timeout_ms=5000):
                        return True
            except Exception:
                continue
        return False

    def go_to_first_page(self):
        prev_selectors = [
            "button[aria-label*='Poprzed']",
            "button[title*='Poprzed']",
            "a[aria-label*='Poprzed']",
            "a[title*='Poprzed']",
            "text=Poprzednia",
            "text=Previous",
            "[aria-label='Go to previous page']",
        ]
        for _ in range(50):
            moved = False
            self.scroll_to_pagination_area()
            before = self.page_signature(self.get_rows_on_page())
            before_page = self.current_numeric_page()
            if before_page is not None and before_page <= 1:
                break
            for sel in prev_selectors:
                try:
                    loc = self.page.locator(sel)
                    if loc.count() == 0:
                        continue
                    for i in range(loc.count()):
                        btn = loc.nth(i)
                        if self.is_disabled_element(btn):
                            continue
                        if not self.click_locator(btn, timeout=2500):
                            continue
                        self.ensure_invoice_list_view()
                        if self.wait_for_page_change(before, before_page, timeout_ms=3500):
                            moved = True
                            break
                    if moved:
                        break
                except Exception:
                    continue
            if not moved:
                if before_page and before_page > 1 and self.click_page_number(1):
                    continue
                break

    def go_to_next_page(self):
        before_signature = self.page_signature(self.get_rows_on_page())
        before_page = self.current_numeric_page()

        if before_page is not None and self.click_page_number(before_page + 1):
            return True

        self.scroll_to_pagination_area()
        next_selectors = [
            "button[aria-label*='Nast']",
            "button[title*='Nast']",
            "a[aria-label*='Nast']",
            "a[title*='Nast']",
            "[role='button'][aria-label*='Nast']",
            "text=Następna",
            "text=Next",
            "text=Dalej",
            "[aria-label='Go to next page']",
        ]
        for sel in next_selectors:
            try:
                loc = self.page.locator(sel)
                if loc.count() == 0:
                    continue
                for i in range(loc.count()):
                    btn = loc.nth(i)
                    if self.is_disabled_element(btn):
                        continue
                    if not self.click_locator(btn, timeout=2500):
                        continue
                    self.ensure_invoice_list_view()
                    if self.wait_for_page_change(before_signature, before_page, timeout_ms=5000):
                        return True
            except Exception:
                continue
        return False

    def scan_all_pages(self):
        all_rows = []
        seen_signatures = set()
        self.go_to_first_page()

        page_no = 0
        max_pages = 500
        while page_no < max_pages:
            page_no += 1
            self.ensure_invoice_list_view()
            self.set_status(f"Skanuję stronę {page_no}...")
            rows = self.get_rows_on_page()
            sig = self.page_signature(rows)

            if sig in seen_signatures:
                self.log(f"[INFO] Wykryto powtórzenie strony przy stronie {page_no}. Kończę skanowanie.")
                break
            seen_signatures.add(sig)

            all_rows.extend(rows)
            self.pages_var.set(str(page_no))
            self.rows_var.set(str(len(all_rows)))
            self.log(f"[INFO] Strona {page_no}: odczytano {len(rows)} wierszy.")
            self.set_progress(page_no, max(page_no + 1, 2))

            if not self.go_to_next_page():
                self.log(f"[INFO] Nie znaleziono kolejnej strony po stronie {page_no}.")
                break

        if page_no >= max_pages:
            self.log("[INFO] Osiągnięto limit bezpieczeństwa 500 stron.")

        return all_rows

    def autosize(self, ws):
        for col in ws.columns:
            max_len = 0
            letter = get_column_letter(col[0].column)
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 38)

    def save_excel(self, rows):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(self.output_dir, f"KSeF_Arkusz_{timestamp}.xlsx")

        wb = Workbook()
        ws_info = wb.active
        ws_info.title = "Info"
        ws_info.append(["Parametr", "Wartość"])
        ws_info.append(["Data eksportu", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws_info.append(["Liczba stron", self.pages_var.get()])
        ws_info.append(["Liczba wierszy", len(rows)])

        ws = wb.create_sheet("Invoices")
        headers = [
            "Identyfikator sprzedawcy",
            "Nazwa sprzedawcy",
            "Nr KSeF",
            "Nr faktury",
            "Data wystawienia",
            "Data zapisania w KSeF",
            "Data otrzymania",
            "Waluta",
            "Netto",
            "Brutto",
            "VAT (PLN)",
        ]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="C81F25")
        header_font = Font(color="FFFFFF", bold=True)

        for idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.fill = header_fill
            cell.font = header_font

        for row in rows:
            ws.append([row.get(h, "") for h in headers])

        self.autosize(ws_info)
        self.autosize(ws)
        wb.save(path)
        return path

    def export_summary(self):
        if self.page is None:
            messagebox.showwarning("Uwaga", "Najpierw kliknij „Otwórz KSeF”.")
            return

        try:
            self.set_busy(True, "Skanowanie listy...")
            self.progress["value"] = 0
            self.pages_var.set("0")
            self.rows_var.set("0")
            rows = self.scan_all_pages()

            if not rows:
                self.set_busy(False, "Brak danych")
                self.log("[BŁĄD] Nie znaleziono żadnych wierszy do eksportu.")
                messagebox.showwarning("Brak danych", "Nie znaleziono żadnych wierszy na liście.")
                return

            self.set_status("Zapisuję plik Excel...")
            self.log("[INFO] Zapisuję plik Excel...")
            self.set_progress(90, 100)
            path = self.save_excel(rows)
            self.file_var.set(os.path.basename(path))
            self.set_progress(100, 100)
            self.set_busy(False, "Gotowe")
            self.log(f"[OK] Zapisano: {path}")
            messagebox.showinfo("Sukces", f"Zestawienie zapisane.\n\n{path}")
        except Exception as e:
            self.set_busy(False, "Błąd")
            self.log(f"[BŁĄD] {e}")
            messagebox.showerror("Błąd", f"Wystąpił błąd.\n\n{e}")

    def on_close(self):
        try:
            if self.context:
                self.context.close()
        except Exception:
            pass
        try:
            if self.browser:
                self.browser.close()
        except Exception:
            pass
        try:
            if self.playwright:
                self.playwright.stop()
        except Exception:
            pass
        self.root.destroy()


def main():
    root = tk.Tk()
    app = KsefSimpleSummaryApp(root)
    root.mainloop()


if __name__ == "__main__":
    try:
        main()
    except Exception:
        traceback.print_exc()
        raise
